"""
Builds 3 class-specific Nucleus Content Excel files:
  templates/Nucleus Class 6 Content.xlsx
  templates/Nucleus Class 7 Content.xlsx
  templates/Nucleus Class 8 Content.xlsx

Each file has 2 tabs (MCQ | QA) with auto-filter enabled on all columns.

Columns:
  MCQ: chapter_id | chapter_name | worksheet | level |
       question | option_a | option_b | option_c | option_d | image_url
  QA:  chapter_id | chapter_name | worksheet | level | question | image_url

Worksheet 1 → level = Elementary
Worksheet 2 → level = Advanced

Run: python3 scripts/build-nucleus-excel.py
"""

import os
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Colours ────────────────────────────────────────────────
BLUE_HEADER  = "1A73E8"
WHITE        = "FFFFFF"
FILLED_MATHS = "EAF1FB"
FILLED_SCI   = "E6F4EA"
BLANK_BG     = "FFFDE7"
BORDER_COLOR = "DADCE0"
ELEM_COLOR   = "1A73E8"
ADV_COLOR    = "D93025"
ELEM_BG      = "E8F0FE"
ADV_BG       = "FCE8E6"
IMG_COLOR    = "7B1FA2"
IMG_BG       = "F3E5F5"

def ws_to_level(ws_num):
    return "Elementary" if str(ws_num) == "1" else "Advanced"

def level_color(level):
    return (ELEM_COLOR, ELEM_BG) if level == "Elementary" else (ADV_COLOR, ADV_BG)

# ── Style helpers ──────────────────────────────────────────
def hdr_font():  return Font(bold=True, color=WHITE, size=11)
def hdr_fill():  return PatternFill("solid", fgColor=BLUE_HEADER)
def thin_border():
    s = Side(style="thin", color=BORDER_COLOR)
    return Border(left=s, right=s, top=s, bottom=s)

def style_header_row(ws, col_count):
    for c in range(1, col_count + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = hdr_font()
        cell.fill = hdr_fill()
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border()

def style_cell(cell, fill=None, bold=False, color="202124",
               center_align=False, italic=False):
    cell.font = Font(bold=bold, color=color, size=10, italic=italic)
    if fill:
        cell.fill = fill
    cell.alignment = (Alignment(horizontal="center", vertical="top")
                      if center_align else Alignment(wrap_text=True, vertical="top"))
    cell.border = thin_border()

def row_fill(chapter_id):
    return (PatternFill("solid", fgColor=FILLED_SCI) if "science" in chapter_id
            else PatternFill("solid", fgColor=FILLED_MATHS))

def blank_fill(): return PatternFill("solid", fgColor=BLANK_BG)

def blank_style(cell):
    cell.fill = blank_fill()
    cell.font = Font(color="BDBDBD", size=10, italic=True)
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    cell.border = thin_border()

# ══════════════════════════════════════════════════════════
# PRE-FILLED DATA (Class 6 only)
# MCQ format: [chapter_id, chapter_name, ws_num, question, opt_a..d]
# QA  format: [chapter_id, chapter_name, ws_num, question]
# answer / explanation / type columns intentionally removed
# ══════════════════════════════════════════════════════════

MCQ_DATA = [
    # ── 6-maths-1 WS1 (Elementary) ────────────────────────
    ["6-maths-1","Patterns in Mathematics","1","What is the next number: 1, 4, 9, 16, __?","20","25","24","21"],
    ["6-maths-1","Patterns in Mathematics","1","The 5th triangular number is:","10","12","15","8"],
    ["6-maths-1","Patterns in Mathematics","1","Which is a Fibonacci number?","14","13","11","16"],
    ["6-maths-1","Patterns in Mathematics","1","Digit sums of multiples of 9 are always:","Equal to 9","Divisible by 9","Odd","Even"],
    ["6-maths-1","Patterns in Mathematics","1","Square numbers are also called:","Triangular numbers","Perfect squares","Prime numbers","Fibonacci numbers"],
    ["6-maths-1","Patterns in Mathematics","1","What is the 4th triangular number?","6","8","10","12"],
    ["6-maths-1","Patterns in Mathematics","1","After 1,1,2,3,5,8 in Fibonacci, what comes next?","12","11","13","14"],
    ["6-maths-1","Patterns in Mathematics","1","Triangular arrangement with 4 rows has how many dots?","8","10","12","6"],
    ["6-maths-1","Patterns in Mathematics","1","What is 2n−1 when n=5?","8","9","10","11"],
    ["6-maths-1","Patterns in Mathematics","1","Which is the sequence of square numbers?","1,2,3,4,5","1,3,6,10,15","1,4,9,16,25","2,4,6,8,10"],
    # ── 6-maths-1 WS2 (Advanced / Assertion-Reason) ───────
    ["6-maths-1","Patterns in Mathematics","2","A: Square numbers always end in 0,1,4,5,6,9. R: These are the only possible units digits of n².","Both A&R true, R explains A","Both true, R doesn't explain A","A true, R false","Both false"],
    ["6-maths-1","Patterns in Mathematics","2","A: 0 is a triangular number. R: T0 = 0×1/2 = 0.","Both A&R true, R explains A","A false, R true","A true, R false","Both false"],
    ["6-maths-1","Patterns in Mathematics","2","A: Sum of first n odd numbers = n². R: 1+3+5+...+(2n-1) = n².","Both A&R true, R explains A","Both true, R doesn't explain A","A true, R false","Both false"],
    ["6-maths-1","Patterns in Mathematics","2","A: Pascal's triangle has only odd numbers. R: Each entry is sum of two previous entries.","A false, R true","Both true","Both false","A true, R false"],
    ["6-maths-1","Patterns in Mathematics","2","A: Fibonacci numbers grow by ratio ≈1.618. R: This ratio is the Golden Ratio.","Both A&R true, R explains A","A true, R false","A false, R true","Both false"],
    ["6-maths-1","Patterns in Mathematics","2","A: Every even number is a triangular number. R: Triangular numbers include 6, 10, 28.","A false, R true","Both true","A true, R false","Both false"],
    ["6-maths-1","Patterns in Mathematics","2","A: Sequence 2,6,12,20,30 is related to triangular numbers. R: Each term = n(n+1).","Both A&R true, R explains A","A false, R true","A true, R false","Both false"],
    ["6-maths-1","Patterns in Mathematics","2","A: 100 is both a perfect square and a triangular number. R: T13=91, T14=105.","A false, R true","Both true, R explains A","A true, R false","Both false"],
    ["6-maths-1","Patterns in Mathematics","2","A: In 1,4,9,16 each term differs from next by consecutive odd numbers. R: 4-1=3, 9-4=5, 16-9=7.","Both A&R true, R explains A","A false, R true","A true, R false","Both false"],
    ["6-maths-1","Patterns in Mathematics","2","A: The sequence 2,4,8,16 is a Fibonacci sequence. R: Each term doubles.","A false, R true","Both true","A true, R false","Both false"],
    # ── 6-maths-2 WS1 ─────────────────────────────────────
    ["6-maths-2","Lines and Angles","1","An angle measuring 145° is called:","Acute","Right","Obtuse","Reflex"],
    ["6-maths-2","Lines and Angles","1","The complement of 37° is:","143°","53°","63°","47°"],
    ["6-maths-2","Lines and Angles","1","Two lines that never meet are called:","Perpendicular","Intersecting","Parallel","Concurrent"],
    ["6-maths-2","Lines and Angles","1","Vertically opposite angles are always:","Supplementary","Complementary","Equal","Adjacent"],
    ["6-maths-2","Lines and Angles","1","The supplement of 75° is:","15°","105°","285°","95°"],
    ["6-maths-2","Lines and Angles","1","A ray has:","Two endpoints","One endpoint","No endpoints","Three endpoints"],
    ["6-maths-2","Lines and Angles","1","A straight angle measures:","90°","270°","360°","180°"],
    ["6-maths-2","Lines and Angles","1","Alternate interior angles on parallel lines are:","Supplementary","Complementary","Equal","Reflex"],
    ["6-maths-2","Lines and Angles","1","A linear pair of angles always sums to:","90°","180°","270°","360°"],
    ["6-maths-2","Lines and Angles","1","Which is NOT a type of angle?","Acute","Obtuse","Diagonal","Reflex"],
    # ── 6-maths-2 WS2 ─────────────────────────────────────
    ["6-maths-2","Lines and Angles","2","A: A right angle is 90°. R: Formed when two perpendicular lines meet.","Both A&R true, R explains A","Both true, R doesn't explain A","A true, R false","A false, R true"],
    ["6-maths-2","Lines and Angles","2","A: Complementary angles must be adjacent. R: Their sum = 90°.","Both A&R true, R explains A","A false, R true","Both false","A true, R false"],
    ["6-maths-2","Lines and Angles","2","A: Vertically opposite angles are supplementary. R: Formed by two intersecting lines.","Both true, R explains A","A false, R true","Both false","A true, R false"],
    ["6-maths-2","Lines and Angles","2","A: A reflex angle is between 180° and 360°. R: Angles greater than a straight angle are reflex.","Both A&R true, R explains A","A true, R false","A false, R true","Both false"],
    ["6-maths-2","Lines and Angles","2","A: Two parallel lines can meet at infinity. R: Parallel lines have no common point.","A true, R false","A false, R true","Both true","Both false"],
    ["6-maths-2","Lines and Angles","2","A: Co-interior angles on parallel lines are supplementary. R: They are on the same side of the transversal.","Both A&R true, R explains A","Both true, R doesn't explain A","A true, R false","A false, R true"],
    ["6-maths-2","Lines and Angles","2","A: All right angles are equal. R: Every right angle measures 90°.","Both A&R true, R explains A","A true, R false","A false, R true","Both false"],
    ["6-maths-2","Lines and Angles","2","A: A straight angle is also a reflex angle. R: 180° is less than 360°.","Both true, R explains A","A false, R true","Both false","A true, R false"],
    ["6-maths-2","Lines and Angles","2","A: Infinitely many lines pass through one point. R: Through one point, any number of lines can be drawn.","Both A&R true, R explains A","Both true, R doesn't explain A","A true, R false","A false, R true"],
    ["6-maths-2","Lines and Angles","2","A: Corresponding angles are equal when lines are parallel. R: They are on same side in matching positions.","Both A&R true, R explains A","A true, R false","A false, R true","Both false"],
    # ── 6-maths-3 WS1 ─────────────────────────────────────
    ["6-maths-3","Number Play","1","Place value of 7 in 47,835?","7","700","7000","70"],
    ["6-maths-3","Number Play","1","Which number is a palindrome?","1234","1221","1342","1023"],
    ["6-maths-3","Number Play","1","Round 6,847 to the nearest thousand:","6000","7000","6800","6900"],
    ["6-maths-3","Number Play","1","Kaprekar's constant for 4-digit numbers:","1729","6174","9801","4321"],
    ["6-maths-3","Number Play","1","Which is an Armstrong number?","123","153","200","321"],
    ["6-maths-3","Number Play","1","A number is divisible by 9 if:","It is even","Ends in 0","Digit sum divisible by 9","Divisible by 3"],
    ["6-maths-3","Number Play","1","Face value of 6 in 76,432?","6000","600","60","6"],
    ["6-maths-3","Number Play","1","Which is NOT a divisibility rule for 5?","Ends in 0","Ends in 5","Digit sum div by 5","Both A and B are correct"],
    ["6-maths-3","Number Play","1","Round 3.74 to the nearest whole number:","3","4","3.7","3.8"],
    ["6-maths-3","Number Play","1","After 57+75=132, what is the next step in the palindrome trick?","132+231=363","132+123=255","132+321=453","132+213=345"],
    # ── 6-maths-3 WS2 ─────────────────────────────────────
    ["6-maths-3","Number Play","2","A: 1729 is smallest number expressible as sum of two cubes in two ways. R: 1729 = 1³+12³ = 9³+10³.","Both A&R true, R explains A","Both true, R doesn't explain A","A true, R false","A false, R true"],
    ["6-maths-3","Number Play","2","A: Rounding always gives the exact answer. R: Rounded numbers are approximations.","A true, R false","A false, R true","Both true","Both false"],
    ["6-maths-3","Number Play","2","A: 371 is an Armstrong number. R: 3³+7³+1³ = 27+343+1 = 371.","Both A&R true, R explains A","A false, R true","A true, R false","Both false"],
    ["6-maths-3","Number Play","2","A: Every palindrome is divisible by 11. R: 121 ÷ 11 = 11.","A false, R true","Both true","A true, R false","Both false"],
    ["6-maths-3","Number Play","2","A: Estimated sum of 248 and 352 is 600. R: 248≈250 and 352≈350; 250+350=600.","Both A&R true, R explains A","A true, R false","A false, R true","Both false"],
    ["6-maths-3","Number Play","2","A: 0 is neither odd nor even. R: 0 is divisible by 2, giving remainder 0.","A false, R true","Both true, R explains A","A true, R false","Both false"],
    ["6-maths-3","Number Play","2","A: 100 is a perfect square. R: 100 = 10 × 10.","Both A&R true, R explains A","Both true, R doesn't explain A","A true, R false","A false, R true"],
    ["6-maths-3","Number Play","2","A: Place value and face value of 0 are always the same. R: Zero × any positional value = 0.","Both A&R true, R explains A","A false, R true","A true, R false","Both false"],
    ["6-maths-3","Number Play","2","A: Kaprekar's operation on 1111 never reaches 6174. R: 1111 has all identical digits — gives 0000.","Both A&R true, R explains A","A true, R false","A false, R true","Both false"],
    ["6-maths-3","Number Play","2","A: A 3-digit palindrome is always divisible by 11. R: ABA = 101A+10B, always divisible by 11.","A false, R true","Both true, R explains A","A true, R false","Both false"],
]

QA_DATA = [
    # ── 6-maths-1 WS1 (Elementary) ────────────────────────
    ["6-maths-1","Patterns in Mathematics","1","What is a pattern? Give two examples from daily life."],
    ["6-maths-1","Patterns in Mathematics","1","Write the first 6 triangular numbers and show how they are formed."],
    ["6-maths-1","Patterns in Mathematics","1","Define square numbers. Find the 7th square number."],
    ["6-maths-1","Patterns in Mathematics","1","What is the Fibonacci sequence? Write its first 10 terms."],
    ["6-maths-1","Patterns in Mathematics","1","How are patterns used in everyday life? Give three examples."],
    # ── 6-maths-1 WS2 (Advanced) ──────────────────────────
    ["6-maths-1","Patterns in Mathematics","2","What is Pascal's Triangle? Write the first 5 rows and explain one pattern you notice."],
    ["6-maths-1","Patterns in Mathematics","2","Explain why the digit sum of every multiple of 9 is divisible by 9. Use three examples."],
    ["6-maths-1","Patterns in Mathematics","2","Create your own pattern with at least 8 terms. State the rule and find the 12th term."],
    ["6-maths-1","Patterns in Mathematics","2","Prove that the sum of first n odd numbers equals n²."],
    ["6-maths-1","Patterns in Mathematics","2","Describe three real-world applications of mathematical patterns."],
    # ── 6-maths-2 WS1 ─────────────────────────────────────
    ["6-maths-2","Lines and Angles","1","Define a line segment and a ray. How are they different from a line?"],
    ["6-maths-2","Lines and Angles","1","What are complementary and supplementary angles? Give one example of each."],
    ["6-maths-2","Lines and Angles","1","Find x if angles (3x+10)° and (2x+5)° form a linear pair."],
    ["6-maths-2","Lines and Angles","1","What are vertically opposite angles? Are they always equal?"],
    ["6-maths-2","Lines and Angles","1","State three angle relationships when a transversal cuts two parallel lines."],
    # ── 6-maths-2 WS2 ─────────────────────────────────────
    ["6-maths-2","Lines and Angles","2","Explain the angles formed when a transversal cuts two parallel lines."],
    ["6-maths-2","Lines and Angles","2","Triangle angles are in ratio 2:3:5. Find each angle and identify the triangle type."],
    ["6-maths-2","Lines and Angles","2","Give three real-life examples each of parallel lines and perpendicular lines."],
    ["6-maths-2","Lines and Angles","2","Two adjacent angles form a straight line. One angle is 3 times the other. Find both angles."],
    ["6-maths-2","Lines and Angles","2","Write two properties each of a line, ray, and line segment."],
    # ── 6-maths-3 WS1 ─────────────────────────────────────
    ["6-maths-3","Number Play","1","What is place value? Find place value and face value of 8 in 38,542."],
    ["6-maths-3","Number Play","1","What is Kaprekar's constant? Show steps for any 4-digit number."],
    ["6-maths-3","Number Play","1","Verify that 153 is an Armstrong number. Also check if 123 is."],
    ["6-maths-3","Number Play","1","Round 47,836 to nearest (a) ten, (b) hundred, (c) thousand."],
    ["6-maths-3","Number Play","1","Check if 8,142 reaches 6174 using Kaprekar's operation (show 2 steps)."],
    # ── 6-maths-3 WS2 ─────────────────────────────────────
    ["6-maths-3","Number Play","2","Explain how rounding is used in everyday life. When is overestimation better?"],
    ["6-maths-3","Number Play","2","Why is 0 considered an even number? Explain using the divisibility rule."],
    ["6-maths-3","Number Play","2","Write the palindrome trick for 89. Show at least 3 steps."],
    ["6-maths-3","Number Play","2","State and explain four divisibility rules with examples."],
    ["6-maths-3","Number Play","2","Explain the difference between estimation and approximation with one example each."],
]

# ══════════════════════════════════════════════════════════
# CHAPTER LIST — read from curriculum.ts at runtime
# ══════════════════════════════════════════════════════════

PREFILLED_IDS = {"6-maths-1", "6-maths-2", "6-maths-3"}
EXCLUDED_IDS  = {"6-science-1", "7-science-1", "8-science-1"}

def load_chapters_from_curriculum():
    import re
    curriculum_path = os.path.join(os.path.dirname(__file__), "..", "data", "curriculum.ts")
    with open(curriculum_path, encoding="utf-8") as f:
        content = f.read()
    chapters = []
    for line in content.splitlines():
        id_m    = re.search(r'id:\s*"([^"]+)"', line)
        title_m = re.search(r'title:\s*"([^"]+)"', line)
        if id_m and title_m:
            chapters.append((id_m.group(1), title_m.group(1)))
    return chapters

ALL_CHAPTERS = load_chapters_from_curriculum()

def get_remaining_chapters(class_num):
    prefix = f"{class_num}-"
    return [(cid, title) for cid, title in ALL_CHAPTERS
            if cid.startswith(prefix)
            and cid not in PREFILLED_IDS
            and cid not in EXCLUDED_IDS]

MCQ_HEADERS = ["chapter_id","chapter_name","worksheet","level",
               "question","option_a","option_b","option_c","option_d","image_url"]
QA_HEADERS  = ["chapter_id","chapter_name","worksheet","level","question","image_url"]

# ══════════════════════════════════════════════════════════
# BUILD SHEETS
# ══════════════════════════════════════════════════════════

def build_mcq_sheet(wb, class_num):
    ws = wb.create_sheet("MCQ")
    headers = MCQ_HEADERS
    ws.append(headers)
    style_header_row(ws, len(headers))
    ws.cell(row=1, column=10).fill = PatternFill("solid", fgColor=IMG_COLOR)
    ws.row_dimensions[1].height = 22

    if class_num == "6":
        for row_data in MCQ_DATA:
            cid, cname, ws_num = row_data[0], row_data[1], row_data[2]
            level = ws_to_level(ws_num)
            full_row = [cid, cname, ws_num, level] + row_data[3:] + [""]
            ws.append(full_row)
            r = ws.max_row
            ws.row_dimensions[r].height = 55
            fill = row_fill(cid)
            fg, bg = level_color(level)
            for c in range(1, len(headers) + 1):
                cell = ws.cell(row=r, column=c)
                if c == 4:
                    style_cell(cell, fill=PatternFill("solid", fgColor=bg),
                               bold=True, color=fg, center_align=True)
                elif c == 3:
                    style_cell(cell, fill=fill, center_align=True)
                elif c == 10:
                    style_cell(cell, fill=PatternFill("solid", fgColor=IMG_BG),
                               color=IMG_COLOR, italic=True)
                else:
                    style_cell(cell, fill=fill)

    for chid, chname in get_remaining_chapters(class_num):
        for ws_num in ["1", "2"]:
            level = ws_to_level(ws_num)
            fg, bg = level_color(level)
            for _ in range(10):
                ws.append([chid, chname, ws_num, level, "", "", "", "", "", ""])
                r = ws.max_row
                ws.row_dimensions[r].height = 40
                for c in range(1, len(headers) + 1):
                    cell = ws.cell(row=r, column=c)
                    if c == 1:
                        style_cell(cell, fill=row_fill(chid), bold=True, color="666666")
                    elif c == 2:
                        style_cell(cell, fill=row_fill(chid), color="666666")
                    elif c == 3:
                        style_cell(cell, fill=row_fill(chid), color="666666", center_align=True)
                    elif c == 4:
                        style_cell(cell, fill=PatternFill("solid", fgColor=bg),
                                   bold=True, color=fg, center_align=True)
                    else:
                        blank_style(cell)

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 44
    ws.column_dimensions["F"].width = 20
    ws.column_dimensions["G"].width = 20
    ws.column_dimensions["H"].width = 20
    ws.column_dimensions["I"].width = 20
    ws.column_dimensions["J"].width = 36
    ws.freeze_panes = "E2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"


def build_qa_sheet(wb, class_num):
    ws = wb.create_sheet("QA")
    headers = QA_HEADERS
    ws.append(headers)
    style_header_row(ws, len(headers))
    ws.cell(row=1, column=6).fill = PatternFill("solid", fgColor=IMG_COLOR)
    ws.row_dimensions[1].height = 22

    if class_num == "6":
        for row_data in QA_DATA:
            cid, cname, ws_num = row_data[0], row_data[1], row_data[2]
            level = ws_to_level(ws_num)
            full_row = [cid, cname, ws_num, level, row_data[3], ""]
            ws.append(full_row)
            r = ws.max_row
            ws.row_dimensions[r].height = 60
            fill = row_fill(cid)
            fg, bg = level_color(level)
            for c in range(1, len(headers) + 1):
                cell = ws.cell(row=r, column=c)
                if c == 4:
                    style_cell(cell, fill=PatternFill("solid", fgColor=bg),
                               bold=True, color=fg, center_align=True)
                elif c == 3:
                    style_cell(cell, fill=fill, center_align=True)
                elif c == 6:
                    style_cell(cell, fill=PatternFill("solid", fgColor=IMG_BG),
                               color=IMG_COLOR, italic=True)
                else:
                    style_cell(cell, fill=fill)

    for chid, chname in get_remaining_chapters(class_num):
        for ws_num in ["1", "2"]:
            level = ws_to_level(ws_num)
            fg, bg = level_color(level)
            for _ in range(5):
                ws.append([chid, chname, ws_num, level, "", ""])
                r = ws.max_row
                ws.row_dimensions[r].height = 50
                for c in range(1, len(headers) + 1):
                    cell = ws.cell(row=r, column=c)
                    if c == 1:
                        style_cell(cell, fill=row_fill(chid), bold=True, color="666666")
                    elif c == 2:
                        style_cell(cell, fill=row_fill(chid), color="666666")
                    elif c == 3:
                        style_cell(cell, fill=row_fill(chid), color="666666", center_align=True)
                    elif c == 4:
                        style_cell(cell, fill=PatternFill("solid", fgColor=bg),
                                   bold=True, color=fg, center_align=True)
                    else:
                        blank_style(cell)

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 50
    ws.column_dimensions["F"].width = 36
    ws.freeze_panes = "E2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"


def main():
    os.makedirs("templates", exist_ok=True)
    for class_num in ["6", "7", "8"]:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        print(f"Building Class {class_num} MCQ tab...")
        build_mcq_sheet(wb, class_num)
        print(f"Building Class {class_num} QA tab...")
        build_qa_sheet(wb, class_num)

        out_path = f"templates/Nucleus Class {class_num} Content.xlsx"
        wb.save(out_path)
        size_kb = os.path.getsize(out_path) // 1024
        remaining = get_remaining_chapters(class_num)
        prefilled = 3 if class_num == "6" else 0
        total = prefilled + len(remaining)
        print(f"  ✅  Saved → {out_path}  ({size_kb} KB)  [{total} chapters]")

    print("\nAll 3 class files generated!")
    print("  Columns: chapter_id | chapter_name | worksheet | level | question | ... | image_url")
    print("  Auto-filter: ON (all columns have dropdown arrows)")
    print("")
    print("  Filter tips (Google Sheets):")
    print("  → Click chapter_id dropdown → pick one chapter → Data → Filter views → Save as filter view")
    print("  → Name it e.g. 'Class 6 - Maths Ch 3 - Elementary'")
    print("  → Switch between saved views from the Filter views menu instantly")


if __name__ == "__main__":
    main()
